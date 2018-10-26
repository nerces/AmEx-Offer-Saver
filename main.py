# file: main.py
"""
This program scrapes the current offers that are available on the AmEx website and then imports an 
xlsx file and uses the data in it to register multiple AmEx cards to the user's credit statement offer of choice.
"""

import requests
from bs4 import BeautifulSoup
import re
import csv
import openpyxl
import sys

def getOfferList():
	
	#Gets the page contents
	page = requests.get("https://www.americanexpress.com/au/network/shopping.html")
	#Parses contents to make it look better
	soup = BeautifulSoup(page.content, 'html.parser')
	#Initialise offerList - This is where each offer's data-gcorid is saved
	offerList = []
	#Finds all 'a' elements with "offer=" in them
	for elem in soup.find_all('a', href=re.compile("offer=")):
		#Extracts the trackingid from the above elem
		offerList.append(elem['trackingid'])
	return offerList

def getOfferDetails(offerList):
	
	#Initialise offerCodes, offerTitles, and offerDescriptions to save the new final data in
	offerCodes = []
	offerTitles = []
	offerDescriptions = []

	for i in offerList:
		#Adding the offer ID to offerCodes
		offerCodes.append(i)
		#Getting the page contents
		page = requests.get("https://www.americanexpress.com/au/network/shopping/doe-offer-detail.html?offer=" + i)
		#Parses content to make it look better
		soup = BeautifulSoup(page.content, 'html.parser')
		#This for loop extracts the offer title from the page
		for title in soup.find_all("h1", class_="supplier"):
			offerTitles.append(title.text.strip())
		#This for loop extracts the offer description from the page
		for description in soup.find_all("div", class_="offer-details"):
			des = description.find('h2')
			offerDescriptions.append(des.text.strip())
	#detailedOfferList is initialised and the other three list are appended to it to create on list with all the details
	detailedOfferList = []
	detailedOfferList.append(offerCodes)
	detailedOfferList.append(offerTitles)
	detailedOfferList.append(offerDescriptions)

	return detailedOfferList

def selectOffers(detailedOfferList):

	print("The current active offers are:\n")
	#Prints out all the offers and their descriptions
	for i in range(0,len(detailedOfferList[0])):
		print("Offer " + str(i + 1) + ":")
		print(detailedOfferList[1][i])
		print(detailedOfferList[2][i])
		print("\n")
	#Asks user to enter an integer to select the offer they want to register for
	offerNumber = input("Please enter the number of the offer you would like to register your cards for: ")
	#Use offerNumber to get the offer's ID for later use. We minus one from offerNumber because the list starts at 0
	offerID = detailedOfferList[0][int(offerNumber) - 1]

	return offerID

def registerCards(offerID):

	#Loads the Excel workbook
	data = openpyxl.load_workbook('details.xlsx')
	#Sets the active sheet
	sheet = data.active
	#Gets the maximum row in the active sheet
	rowCount = sheet.max_row

	for i in range (2, rowCount):

	    #Details for the form extracted from the Excel sheet
	    firstName = sheet.cell(row = i, column = 1).value
	    lastName = sheet.cell(row = i, column = 2).value
	    cardNumber = sheet.cell(row = i, column = 3).value
	    email = sheet.cell(row = i, column = 4).value

		r = requests.post("https://www.americanexpress.com/gemservices/shopping/enrolment.submit/au/",
			data = {'firstName' : '%s' % firstName,
			'lastName' : '%s' % lastName,
			'email' : '%s' % email,
			'cardNumber' : '%s' % cardNumber,
			'offerId' : offerID})
		if r.status_code != 200:
			print(r.status_code, r.reason)
			sys.exit("Looks like we're having some problems with your internet connection or the AmEx servers.\nNow exiting.")
		else:
			print(r.text[:300])

def main():

	offerList = getOfferList()
	detailedOfferList = getOfferDetails(offerList)
	offerID = selectOffers(detailedOfferList)
	registerCards(offerID)

main()